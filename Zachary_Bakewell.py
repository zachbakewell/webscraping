from urllib.request import urlopen,Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

# scrape the website below to retrieve the top 5 countries with the highest GDPs. Calculate the GDP per capita
# by dividing the GDP by the population. You can perform the calculation in Python natively or insert the code
# in excel that will perform the calculation in Excel by each row. DO NOT scrape the GDP per capita from the
# webpage, make sure you use your own calculation.


### REMEMBER ##### - your output should match the excel file (GDP_Report.xlsx) including all formatting.

webpage = 'https://www.worldometers.info/gdp/gdp-by-country/'

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(webpage, headers=headers)
webpage = urlopen(req).read()
soup = BeautifulSoup(webpage, 'html.parser')

font_1 = Font(name='Calibri', size=24, bold=True, italic=False)
chart_font = Font(name="Calibri Light", size=13)

wb = xl.Workbook()

ws = wb.active

ws.title = 'Country Table'

ws["A1"] = 'No.'
ws["B1"] = "country"
ws["C1"] = "GDP"
ws["D1"] = "Population"
ws["E1"] = "GDP Per Capita"

ws.column_dimensions["A"].width = 10
ws["A1"].font = font_1

ws.column_dimensions["B"].width = 25
ws["B1"].font = font_1

ws.column_dimensions["C"].width = 25
ws["C1"].font = font_1

ws.column_dimensions["D"].width = 25
ws["D1"].font = font_1

ws.column_dimensions["E"].width = 30
ws["E1"].font = font_1

country_list = soup.findAll('tr')

for row in range(1,6):
    td = country_list[row].findAll('td')
    if td:
        no = td[0].text
        country = td[1].text
        gdp = td[2].text.replace("$","")
        gdpCalc = td[2].text.replace("$","").replace(',','')
        populationCalc = td[5].text.replace(',','')
        population = td[5].text
        per_cap = float(gdpCalc)/int(populationCalc)

        formatGDP = str("${:.25}".format(gdp))
        formatCap = str("${:,.2f}".format(per_cap))

        ws['A' + str(row+1)] = no
        ws['B' + str(row+1)] = country
        ws['C' + str(row+1)] = formatGDP
        ws['D' + str(row+1)] = population
        ws['E' + str(row+1)] = formatCap

        ws['A' + str(row+1)].font = chart_font
        ws['B' + str(row+1)].font = chart_font
        ws['C' + str(row+1)].font = chart_font
        ws['D' + str(row+1)].font = chart_font
        ws['E' + str(row+1)].font = chart_font

        for cell in ws['C:C']:
            cell.number_format = u'"$ "#,##0.00'

        for cell in ws['E:E']:
            cell.number_format = u'"$ "#,##0.00'

wb.save("CountryReport.xlsx")