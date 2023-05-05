from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
from twilio.rest import Client
import keys

font_1 = Font(name="Calibri", size=18, color='ECAA11')
chart_font = Font(name="Calibri Light", size=13)

wb = xl.Workbook()

ws = wb.active

ws.title = 'Crypto Table'

ws["A1"] = 'Name'
ws["B1"] = "Symbol"
ws["C1"] = "Current Price"
ws["D1"] = "% Change"
ws["E1"] = "Corresponding Price"

ws.column_dimensions["A"].width = 25
ws["A1"].font = font_1

ws.column_dimensions["B"].width = 25
ws["B1"].font = font_1

ws.column_dimensions["C"].width = 25
ws["C1"].font = font_1

ws.column_dimensions["D"].width = 25
ws["D1"].font = font_1

ws.column_dimensions["E"].width = 25
ws["E1"].font = font_1


webpage = 'https://www.coingecko.com'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url=webpage, headers=headers)

page = urlopen(req)

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

#print(title.text)


client = Client(keys.accountSID, keys.authToken)

TwilioNumber = ''

mycellphone = ''

crypto_list = soup.findAll('tr')

for row in range(1,6):
    td = crypto_list[row].findAll('td')
    if td:
        name_symbol_comb = td[2].text.split()
        name = name_symbol_comb[0]
        symbol = name_symbol_comb[1]
        price = td[3].text
        perc_change = td[5].text.replace('%','')


        price_change = float(perc_change)/100
        price1 = float(price.replace(',','').replace('$',''))

        if price_change < 0:
            new_price = price1 * (1 - price_change)
            old_price = str("${:,.2f}".format(new_price))
        else:
            new_price = float(price1) * (1+ price_change)
            old_price = str("${:,.2f}".format(new_price))

        ws['A' + str(row+1)] = name
        ws['B' + str(row+1)] = symbol
        ws['C' + str(row+1)] = price
        ws['D' + str(row+1)] = (perc_change + '%')
        ws['E' + str(row+1)] = old_price

        ws['A' + str(row+1)].font = chart_font
        ws['B' + str(row+1)].font = chart_font
        ws['C' + str(row+1)].font = chart_font
        ws['D' + str(row+1)].font = chart_font
        ws['E' + str(row+1)].font = chart_font


        if symbol == "ETH" or symbol == "BTC":
            value_check = price1 - new_price
            if value_check > 5:
                textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber, body= f"{symbol} has increased to {price.strip()}!!!")
            if value_check < 5:
                textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber, body= f"{symbol} has decreased to {price.strip()}!!!")


wb.save("CryptoReport.xlsx")